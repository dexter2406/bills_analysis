from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from bills_analysis.excel_ops import (
    normalize_date,
    threshold_for,
    to_score,
    write_datum_cell,
)


def load_results(path: Path) -> list[dict[str, Any]]:
    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        return []
    if raw.lstrip().startswith("["):
        data = json.loads(raw)
        if isinstance(data, list):
            return data
        raise ValueError("Results JSON is not a list.")
    return [json.loads(line) for line in raw.splitlines() if line.strip()]


def load_config(path: Path) -> dict[str, Any]:
    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        raise ValueError(f"Empty config file: {path}")
    data = json.loads(raw)
    if not isinstance(data, dict):
        raise ValueError("Config JSON must be an object.")
    return data


def to_link(value: Any, base_dir: Path) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("http://") or text.startswith("https://"):
        return text
    path = Path(text)
    if not path.is_absolute():
        path = (base_dir / path).resolve()
    try:
        return path.as_uri()
    except ValueError:
        return str(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Map OFFICE results JSON to Excel (one row per entry)."
    )
    parser.add_argument("json_path", type=Path, help="Path to results JSON")
    parser.add_argument(
        "excel_path",
        type=Path,
        nargs="?",
        help="Output Excel path (default: same dir/name as JSON)",
    )
    args = parser.parse_args()

    path = args.json_path
    out_path = args.excel_path or path.with_suffix(".xlsx")
    items = load_results(path)
    config = load_config(Path(__file__).with_name("config.json"))

    headers = [
        "Datum",
        "Type",
        "Rechnung Name",
        "Brutto",
        "Netto",
        "Steuernummer",
        "Is Receiver OK",
        "need review",
        "Rechnung Scannen",
    ]
    rows: list[list[Any]] = []
    row_meta: list[dict[str, Any]] = []

    for item in items:
        category = str(item.get("category") or "").strip().lower()
        if category != "office":
            continue
        result = item.get("result") or {}
        rows.append(
            [
                normalize_date(result.get("run_date")) or result.get("run_date"),
                result.get("type"),
                result.get("sender"),
                result.get("brutto"),
                result.get("netto"),
                result.get("tax_id"),
                result.get("receiver_ok"),
                False,
                result.get("preview_path") or item.get("preview_path"),
            ]
        )
        row_meta.append(
            {
                "brutto": result.get("brutto"),
                "netto": result.get("netto"),
                "total_tax": result.get("total_tax"),
                "receiver_ok": result.get("receiver_ok"),
                "preview_path": result.get("preview_path") or item.get("preview_path"),
                "score": item.get("score") or {},
            }
        )

    if not rows:
        print("No OFFICE rows found.")
        raise SystemExit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Office"
    ws.append(headers)
    orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for row_idx, row in enumerate(rows, start=2):
        ws.append(row)
        # Ensure Datum is stored as date type
        write_datum_cell(ws.cell(row=row_idx, column=1), row[0])
        meta = row_meta[row_idx - 2]
        score = meta.get("score") or {}
        need_review = False
        # brutto/netto/total_tax threshold check
        for field, header in [
            ("brutto", "Brutto"),
            ("netto", "Netto"),
        ]:
            score_val = to_score(score.get(field))
            if score_val is None or score_val < threshold_for(field, config):
                col = headers.index(header) + 1
                ws.cell(row=row_idx, column=col).fill = orange
                need_review = True
        # Steuernummer must not be empty
        tax_val = row[headers.index("Steuernummer")]
        if tax_val in (None, "", "None"):
            col = headers.index("Steuernummer") + 1
            ws.cell(row=row_idx, column=col).fill = orange
            need_review = True
        # receiver_ok must be True
        if meta.get("receiver_ok") is not True:
            col = headers.index("Is Receiver OK") + 1
            ws.cell(row=row_idx, column=col).fill = orange
            need_review = True
        # write need review
        col = headers.index("need review") + 1
        ws.cell(row=row_idx, column=col, value=need_review)
        if need_review:
            ws.cell(row=row_idx, column=col).fill = orange
        # pdf link
        link = to_link(meta.get("preview_path"), path.parent)
        if link:
            col = headers.index("Rechnung Scannen") + 1
            cell = ws.cell(row=row_idx, column=col)
            cell.value = "check pdf"
            cell.hyperlink = link
    wb.save(out_path)
    print(f"[Excel] Written: {out_path}")


if __name__ == "__main__":
    main()
