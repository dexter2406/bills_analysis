from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from bills_analysis.excel_ops import (
    build_rows_with_meta,
    low_confidence_fields,
    normalize_date,
    threshold_for,
    to_score,
)


def load_results(path: Path) -> list[dict[str, Any]]:
    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        return []
    if raw.lstrip().startswith("["):
        data = json.loads(raw)
        if isinstance(data, list):
            return data
        raise ValueError("Results JSON is not a list.")
    return [json.loads(line) for line in raw.splitlines() if line.strip()]


def load_thresholds(path: Path) -> dict[str, Any]:
    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        raise ValueError(f"Empty thresholds file: {path}")
    data = json.loads(raw)
    if not isinstance(data, dict):
        raise ValueError("Thresholds JSON must be an object.")
    return data


def max_pages(thresholds: dict[str, Any], default: int = 4) -> int:
    value = thresholds.get("max_pages", default)
    try:
        value = int(value)
    except (TypeError, ValueError):
        return default
    return value


def log_low_conf(item: dict[str, Any], thresholds: dict[str, Any]) -> None:
    result = item.get("result") or {}
    score = item.get("score") or {}
    low_fields = low_confidence_fields(result, score, thresholds)
    if not low_fields:
        return
    filename = str(item.get("filename") or "")
    for field in sorted(low_fields):
        raw_score = score.get(field)
        score_val = to_score(raw_score)
        if field in {"brutto", "netto"} and score_val == -1:
            raw_score = score.get("total_tax")
            threshold = threshold_for("total_tax", thresholds)
        else:
            threshold = threshold_for(field, thresholds)
        print(
            f"[LOW_CONF] file={filename} field={field} score={raw_score} "
            f"threshold={threshold} reason=below_threshold_or_missing"
        )


def to_link(value: Any, base_dir: Path) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("http://") or text.startswith("https://"):
        return text
    path = Path(text)
    if not path.is_absolute():
        path = (base_dir / path).resolve()
    try:
        return path.as_uri()
    except ValueError:
        return str(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Map results JSON to a one-row Excel file."
    )
    parser.add_argument("json_path", type=Path, help="Path to results JSON")
    parser.add_argument(
        "excel_path",
        type=Path,
        nargs="?",
        help="Output Excel path (default: same dir/name as JSON)",
    )
    args = parser.parse_args()

    path = args.json_path
    out_path = args.excel_path or path.with_suffix(".xlsx")
    items = load_results(path)
    thresholds_path = Path(__file__).with_name("config.json")
    thresholds = load_thresholds(thresholds_path)
    for item in items:
        log_low_conf(item, thresholds)
    rows, zbon_files_by_date = build_rows_with_meta(items, thresholds)
    if not rows:
        print("No rows generated.")
        raise SystemExit(1)
    if len(rows) > 1:
        print(f"[WARN] Multiple dates found ({len(rows)}). Only the first row will be written.")

    first = rows[0]
    # need review = (low confidence) OR (page_count > max_pages)
    datum = first.get("Datum") or "UNKNOWN"
    need_review = bool(first.get("need review"))
    for item in items:
        result = item.get("result") or {}
        run_date = normalize_date(result.get("run_date")) or "UNKNOWN"
        if run_date != datum:
            continue
        page_count = item.get("page_count")
        if isinstance(page_count, int) and page_count > max_pages(thresholds):
            need_review = True
            break
    first["need review"] = need_review
    headers = list(first.keys())

    thresholds = load_thresholds(Path(__file__).with_name("config.json"))
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(headers)
    ws.append([first.get(h) for h in headers])

    # Highlight low-confidence fields in orange
    header_to_col = {name: idx + 1 for idx, name in enumerate(headers)}
    data_row_idx = 2
    # Map low-confidence fields to the Ausgabe slot based on actual output order
    datum = first.get("Datum") or "UNKNOWN"
    low_headers = set()
    # ZBon items (Umsatz)
    for item in items:
        result = item.get("result") or {}
        run_date = normalize_date(result.get("run_date")) or "UNKNOWN"
        if run_date != datum:
            continue
        category = str(item.get("category") or "").strip().lower()
        if category != "zbon":
            continue
        low_fields = low_confidence_fields(result, item.get("score") or {}, thresholds)
        if "brutto" in low_fields:
            low_headers.add("Umsatz Brutto")
        if "netto" in low_fields:
            low_headers.add("Umsatz Netto")
    # BAR items by slot
    zbon_files = zbon_files_by_date.get(datum, [])
    for idx, fname in enumerate(zbon_files, start=1):
        for item in items:
            if str(item.get("filename") or "") != fname:
                continue
            low_fields = low_confidence_fields(
                item.get("result") or {},
                item.get("score") or {},
                thresholds,
            )
            if "store_name" in low_fields:
                low_headers.add(f"Ausgabe {idx} Name")
            if "brutto" in low_fields:
                low_headers.add(f"Ausgabe {idx} Brutto")
            if "netto" in low_fields:
                low_headers.add(f"Ausgabe {idx} Netto")
            break

    for header in low_headers:
        col = header_to_col.get(header)
        if col is not None:
            ws.cell(row=data_row_idx, column=col).fill = orange_fill
    if need_review:
        col = header_to_col.get("need review")
        if col is not None:
            ws.cell(row=data_row_idx, column=col).fill = orange_fill

    # Add preview links under Ausgabe N Name
    preview_map = {}
    for item in items:
        filename = str(item.get("filename") or "")
        preview_path = item.get("preview_path")
        if filename:
            preview_map[filename] = preview_path
    link_row_idx = data_row_idx + 1
    zbon_files = zbon_files_by_date.get(datum, [])
    for idx, fname in enumerate(zbon_files, start=1):
        if idx > 5:
            break
        preview = preview_map.get(fname)
        link = to_link(preview, path.parent) if preview else None
        if not link:
            continue
        col = header_to_col.get(f"Ausgabe {idx} Name")
        if col is None:
            continue
        cell = ws.cell(row=link_row_idx, column=col)
        cell.value = "check pdf"
        cell.hyperlink = link

    wb.save(out_path)
    print(f"[Excel] Written: {out_path}")
    print(json.dumps(rows, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
