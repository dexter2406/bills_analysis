from __future__ import annotations

import argparse
import shutil
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bills_analysis.excel_ops import (
    merge_validated_row,
    normalize_date,
    normalize_header,
    write_datum_cell,
)

def cell_has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def load_single_row(path: Path) -> tuple[list[str], list[Any]]:
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if ws.max_row < 2:
        raise ValueError("Validated Excel must contain exactly one data row.")
    if ws.max_row > 2:
        print(f"[WARN] Validated Excel has {ws.max_row - 1} data rows; only the first will be used.")
    row_values = [cell.value for cell in ws[2]]
    return [str(h).strip() if h is not None else "" for h in headers], row_values


def find_row_by_datum(ws, datum: str) -> int | None:
    target = normalize_date(datum) or str(datum).strip()
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        cell_text = normalize_date(cell_value) or str(cell_value).strip()
        if cell_text == target:
            return row_idx
    return None


def merge_daily(
    validated_xlsx: Path,
    monthly_xlsx: Path,
    *,
    out_dir: Path | None = None,
) -> Path:
    validated_headers, validated_row = load_single_row(validated_xlsx)
    if not validated_headers or validated_headers[0] != "Datum":
        raise ValueError("Validated Excel must have 'Datum' as the first column.")
    datum = str(validated_row[0]).strip() if validated_row else ""
    if not datum:
        raise ValueError("Validated Excel has empty Datum.")

    out_dir = out_dir or monthly_xlsx.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = int(datetime.now().timestamp())
    out_path = out_dir / f"full_result_{timestamp}.xlsx"
    if out_path.resolve() == validated_xlsx.resolve():
        raise ValueError("输出路径与已校验文件相同，请指定不同的输出目录。")
    if out_path.resolve() == monthly_xlsx.resolve():
        raise ValueError("输出路径与全量文件相同，请指定不同的输出目录。")

    shutil.copy2(monthly_xlsx, out_path)
    print(f"[Excel] Copied monthly file -> {out_path}")
    wb = load_workbook(out_path)
    ws = wb.active

    monthly_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    header_to_col = {name: idx + 1 for idx, name in enumerate(monthly_headers)}

    target_row = find_row_by_datum(ws, datum)
    if target_row is None:
        raise ValueError(f"Datum not found in monthly Excel: {datum}")

    # Warning if any existing content in the row (excluding Datum)
    has_content = False
    for col_idx in range(2, ws.max_column + 1):
        if cell_has_value(ws.cell(row=target_row, column=col_idx).value):
            has_content = True
            break
    if has_content:
        print(f"[WARN] Existing content for Datum {datum} will be overwritten.")

    # Write values by matching headers
    updates, missing_headers = merge_validated_row(
        validated_headers,
        validated_row,
        monthly_headers,
    )
    for header, value in updates.items():
        col_idx = header_to_col.get(header)
        if col_idx is None:
            continue
        cell = ws.cell(row=target_row, column=col_idx)
        if normalize_header(header) == normalize_header("Datum"):
            write_datum_cell(cell, value)
        else:
            cell.value = value
    if missing_headers:
        print(f"[WARN] Headers not found in monthly Excel: {missing_headers}")

    wb.save(out_path)
    print(f"[Excel] Merged and saved: {out_path}")
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Merge a validated one-row Excel into a monthly Excel by Datum."
    )
    parser.add_argument("validated_xlsx", type=Path, help="One-row validated Excel")
    parser.add_argument("monthly_xlsx", type=Path, help="Monthly full Excel")
    parser.add_argument(
        "--out-dir",
        dest="out_dir",
        type=Path,
        help="Output directory (default: same as monthly_xlsx)",
    )
    args = parser.parse_args()
    merge_daily(args.validated_xlsx, args.monthly_xlsx, out_dir=args.out_dir)


if __name__ == "__main__":
    main()
