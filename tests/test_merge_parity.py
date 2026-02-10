from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from bills_analysis.services.merge_service import merge_daily, merge_office


def _new_book(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    """Create workbook fixture with headers and data rows."""

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_daily_merge_overwrite_parity() -> None:
    """Daily merge should overwrite matching Datum row fields by header mapping."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_daily.xlsx"
    monthly = root / "monthly_daily.xlsx"

    _new_book(
        validated,
        ["Datum", "Umsatz Brutto", "Umsatz Netto", "need review"],
        [["04/02/2026", 120.0, 100.0, False]],
    )
    _new_book(
        monthly,
        ["Datum", "Umsatz Brutto", "Umsatz Netto"],
        [["04/02/2026", 1.0, 2.0]],
    )

    out_path = merge_daily(validated, monthly, out_dir=root)
    wb = load_workbook(out_path)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == 120.0
    assert ws.cell(row=2, column=3).value == 100.0


def test_office_merge_append_parity() -> None:
    """Office merge append mode should append validated rows and keep hyperlinks."""

    root = Path("outputs") / "pytest_tmp" / str(uuid4())
    validated = root / "validated_office.xlsx"
    monthly = root / "monthly_office.xlsx"

    _new_book(
        validated,
        ["Datum", "Type", "Rechnung Name", "need review", "Rechnung Scannen"],
        [["04/02/2026", "Miete", "Metro", True, "check pdf"]],
    )
    wb_validated = load_workbook(validated)
    ws_validated = wb_validated.active
    ws_validated.cell(row=2, column=5).hyperlink = "file:///tmp/demo.pdf"
    wb_validated.save(validated)

    _new_book(
        monthly,
        ["Datum", "Type", "Rechnung Name", "Rechnung Scannen"],
        [["03/02/2026", "Alt", "Old", None]],
    )

    out_path = merge_office(validated, monthly, out_dir=root, append=True)
    wb = load_workbook(out_path)
    ws = wb.active
    assert ws.max_row == 3
    assert ws.cell(row=3, column=2).value == "Miete"
    assert ws.cell(row=3, column=3).value == "Metro"
