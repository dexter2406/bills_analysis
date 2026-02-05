from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bills_analysis.excel_ops import normalize_date, normalize_header, write_datum_cell


def _cell_has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _load_all_rows(path: Path) -> tuple[list[str], list[list[Any]], list[list[str | None]]]:
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows: list[list[Any]] = []
    links: list[list[str | None]] = []
    for row_idx in range(2, ws.max_row + 1):
        row_cells = ws[row_idx]
        rows.append([cell.value for cell in row_cells])
        links.append(
            [cell.hyperlink.target if cell.hyperlink is not None else None for cell in row_cells]
        )
    return [str(h).strip() if h is not None else "" for h in headers], rows, links


def merge_office(
    validated_xlsx: Path,
    monthly_xlsx: Path,
    *,
    out_dir: Path | None = None,
    append: bool = False,
) -> Path:
    validated_headers, validated_rows, validated_links = _load_all_rows(validated_xlsx)
    if not validated_headers or normalize_header(validated_headers[0]) != normalize_header("Datum"):
        raise ValueError("Validated Excel must have 'Datum' as the first column.")
    if not validated_rows:
        raise ValueError("Validated Excel has no data rows.")

    out_dir = out_dir or monthly_xlsx.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = int(datetime.now().timestamp())
    out_path = out_dir / f"full_result_{timestamp}.xlsx"
    if out_path.resolve() == validated_xlsx.resolve():
        raise ValueError("输出路径与已校验文件相同，请指定不同的输出目录。")
    if out_path.resolve() == monthly_xlsx.resolve():
        raise ValueError("输出路径与全量文件相同，请指定不同的输出目录。")

    shutil.copy2(monthly_xlsx, out_path)
    print(f"[Excel] Copied monthly file -> {out_path}")
    wb = load_workbook(out_path)
    ws = wb.active

    monthly_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

    # Determine target row for overwrite mode (by Datum)
    def _find_row_by_datum(datum: Any) -> int | None:
        target = normalize_date(datum) or str(datum).strip()
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            cell_norm = normalize_date(cell_value) or str(cell_value).strip()
            if cell_norm == target:
                return row_idx
        return None

    for row, row_links in zip(validated_rows, validated_links):
        # remove "need review" column if present (case-insensitive)
        if any(normalize_header(h) == normalize_header("need review") for h in validated_headers):
            filtered_headers = []
            filtered_row = []
            filtered_links: list[str | None] = []
            for h, v in zip(validated_headers, row):
                if normalize_header(h) == normalize_header("need review"):
                    continue
                filtered_headers.append(h)
                filtered_row.append(v)
            for h, link in zip(validated_headers, row_links):
                if normalize_header(h) == normalize_header("need review"):
                    continue
                filtered_links.append(link)
        else:
            filtered_headers = validated_headers
            filtered_row = row
            filtered_links = row_links

        if append:
            ws.append(filtered_row)
            row_idx = ws.max_row
            if filtered_row:
                write_datum_cell(ws.cell(row=row_idx, column=1), filtered_row[0])
            for col_idx, link in enumerate(filtered_links, start=1):
                if link:
                    ws.cell(row=row_idx, column=col_idx).hyperlink = link
            continue

        datum_val = filtered_row[0] if filtered_row else ""
        target_row = _find_row_by_datum(datum_val)
        if target_row is None:
            # If not found, append new row
            ws.append(filtered_row)
            continue

        # Warn if existing content
        has_content = False
        for col_idx in range(2, ws.max_column + 1):
            if _cell_has_value(ws.cell(row=target_row, column=col_idx).value):
                has_content = True
                break
        if has_content:
            print(f"[WARN] Existing content for Datum {datum_val} will be overwritten.")

        # Direct row copy by position (no header mapping)
        for col_idx, (value, link) in enumerate(zip(filtered_row, filtered_links), start=1):
            if col_idx == 1:
                write_datum_cell(ws.cell(row=target_row, column=col_idx), value)
                continue
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            cell.hyperlink = link

    wb.save(out_path)
    print(f"[Excel] Merged and saved: {out_path}")
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Merge office validated Excel into monthly Excel."
    )
    parser.add_argument("validated_xlsx", type=Path, help="Validated Excel")
    parser.add_argument("monthly_xlsx", type=Path, help="Monthly Excel")
    parser.add_argument(
        "--out-dir",
        dest="out_dir",
        type=Path,
        help="Output directory (default: same as monthly_xlsx)",
    )
    parser.add_argument(
        "--append",
        action="store_true",
        help="Append rows instead of overwrite by Datum",
    )
    args = parser.parse_args()
    merge_office(args.validated_xlsx, args.monthly_xlsx, out_dir=args.out_dir, append=args.append)


if __name__ == "__main__":
    main()
