from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bills_analysis.excel_ops import (
    merge_validated_row,
    normalize_date,
    normalize_header,
    write_datum_cell,
)


def _cell_has_value(value: Any) -> bool:
    """Check whether an Excel cell value should be treated as non-empty."""

    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _load_single_row(path: Path) -> tuple[list[str], list[Any]]:
    """Load first data row and headers from validated workbook."""

    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if ws.max_row < 2:
        raise ValueError("Validated Excel must contain exactly one data row.")
    row_values = [cell.value for cell in ws[2]]
    return [str(h).strip() if h is not None else "" for h in headers], row_values


def _find_row_by_datum(ws: Any, datum: str) -> int | None:
    """Find monthly workbook row index by normalized Datum value."""

    target = normalize_date(datum) or str(datum).strip()
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        cell_text = normalize_date(cell_value) or str(cell_value).strip()
        if cell_text == target:
            return row_idx
    return None


def merge_daily_excel(
    validated_xlsx: Path,
    monthly_xlsx: Path,
    *,
    out_dir: Path | None = None,
) -> Path:
    """Merge daily validated one-row Excel into monthly workbook by Datum."""

    validated_headers, validated_row = _load_single_row(validated_xlsx)
    if not validated_headers or validated_headers[0] != "Datum":
        raise ValueError("Validated Excel must have 'Datum' as the first column.")
    datum = str(validated_row[0]).strip() if validated_row else ""
    if not datum:
        raise ValueError("Validated Excel has empty Datum.")

    out_dir = out_dir or monthly_xlsx.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = int(datetime.now().timestamp())
    out_path = out_dir / f"full_result_{timestamp}.xlsx"
    if out_path.resolve() == validated_xlsx.resolve():
        raise ValueError("输出路径与已校验文件相同，请指定不同的输出目录。")
    if out_path.resolve() == monthly_xlsx.resolve():
        raise ValueError("输出路径与全量文件相同，请指定不同的输出目录。")

    shutil.copy2(monthly_xlsx, out_path)
    wb = load_workbook(out_path)
    ws = wb.active
    monthly_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    header_to_col = {name: idx + 1 for idx, name in enumerate(monthly_headers)}

    target_row = _find_row_by_datum(ws, datum)
    if target_row is None:
        raise ValueError(f"Datum not found in monthly Excel: {datum}")

    updates, _missing_headers = merge_validated_row(validated_headers, validated_row, monthly_headers)
    for header, value in updates.items():
        col_idx = header_to_col.get(header)
        if col_idx is None:
            continue
        cell = ws.cell(row=target_row, column=col_idx)
        if normalize_header(header) == normalize_header("Datum"):
            write_datum_cell(cell, value)
        else:
            cell.value = value

    wb.save(out_path)
    return out_path


def _load_all_rows(path: Path) -> tuple[list[str], list[list[Any]], list[list[str | None]]]:
    """Load all data rows and hyperlink targets from validated workbook."""

    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows: list[list[Any]] = []
    links: list[list[str | None]] = []
    for row_idx in range(2, ws.max_row + 1):
        row_cells = ws[row_idx]
        rows.append([cell.value for cell in row_cells])
        links.append([cell.hyperlink.target if cell.hyperlink is not None else None for cell in row_cells])
    return [str(h).strip() if h is not None else "" for h in headers], rows, links


def merge_office_excel(
    validated_xlsx: Path,
    monthly_xlsx: Path,
    *,
    out_dir: Path | None = None,
    append: bool = False,
) -> Path:
    """Merge office validated Excel into monthly workbook in overwrite/append mode."""

    validated_headers, validated_rows, validated_links = _load_all_rows(validated_xlsx)
    if not validated_headers or normalize_header(validated_headers[0]) != normalize_header("Datum"):
        raise ValueError("Validated Excel must have 'Datum' as the first column.")
    if not validated_rows:
        raise ValueError("Validated Excel has no data rows.")

    out_dir = out_dir or monthly_xlsx.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = int(datetime.now().timestamp())
    out_path = out_dir / f"full_result_{timestamp}.xlsx"
    if out_path.resolve() == validated_xlsx.resolve():
        raise ValueError("输出路径与已校验文件相同，请指定不同的输出目录。")
    if out_path.resolve() == monthly_xlsx.resolve():
        raise ValueError("输出路径与全量文件相同，请指定不同的输出目录。")

    shutil.copy2(monthly_xlsx, out_path)
    wb = load_workbook(out_path)
    ws = wb.active

    def _find_row_by_datum(datum: Any) -> int | None:
        """Resolve target row by Datum value for overwrite mode."""

        target = normalize_date(datum) or str(datum).strip()
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            cell_norm = normalize_date(cell_value) or str(cell_value).strip()
            if cell_norm == target:
                return row_idx
        return None

    for row, row_links in zip(validated_rows, validated_links):
        if any(normalize_header(h) == normalize_header("need review") for h in validated_headers):
            filtered_headers = []
            filtered_row = []
            filtered_links: list[str | None] = []
            for h, v in zip(validated_headers, row):
                if normalize_header(h) == normalize_header("need review"):
                    continue
                filtered_headers.append(h)
                filtered_row.append(v)
            for h, link in zip(validated_headers, row_links):
                if normalize_header(h) == normalize_header("need review"):
                    continue
                filtered_links.append(link)
        else:
            filtered_headers = validated_headers
            filtered_row = row
            filtered_links = row_links

        if append:
            ws.append(filtered_row)
            row_idx = ws.max_row
            if filtered_row:
                write_datum_cell(ws.cell(row=row_idx, column=1), filtered_row[0])
            for col_idx, link in enumerate(filtered_links, start=1):
                if link:
                    ws.cell(row=row_idx, column=col_idx).hyperlink = link
            continue

        datum_val = filtered_row[0] if filtered_row else ""
        target_row = _find_row_by_datum(datum_val)
        if target_row is None:
            ws.append(filtered_row)
            continue

        for col_idx, (value, link) in enumerate(zip(filtered_row, filtered_links), start=1):
            if col_idx == 1:
                write_datum_cell(ws.cell(row=target_row, column=col_idx), value)
                continue
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            cell.hyperlink = link

    wb.save(out_path)
    return out_path
